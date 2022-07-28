from zipfile import ZipFile
import os
from os.path import basename
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook


def create_zip_archive(dir_name):
    with ZipFile('../zip_helpers/resources/my_archive.zip', 'w') as zip_:
        for folder_name, subfolders, filenames in os.walk(dir_name):
            for filename in filenames:
                file_path = os.path.join(folder_name, filename)
                zip_.write(file_path, basename(file_path))
    zip_.close()
    return zip_


def test_files():
    curr_dir = "../zip_helpers/files/"
    zip_file = create_zip_archive(curr_dir)
    for file in zip_file.namelist():
        if file.endswith('.pdf'):
            pdf = PdfReader(curr_dir + file)
            assert len(pdf.pages) == 408, f'number of pages should be 408, got {len(pdf.pages)} instead'
        elif file.endswith('.xlsx'):
            xlsx = load_workbook(curr_dir + file)
            sheet = xlsx.active
            assert sheet.cell(7,
                              3).value == "Brumm", f'Value should be equal "Brumm", got {sheet.cell(7, 3).value} instead'
        elif file.endswith('.csv'):
            csv_reader = csv.reader(curr_dir + file)
            row_count = sum(1 for row in csv_reader)
            assert row_count == 34, f"Count of rows must be 34, got {row_count}, instead"


test_files()
